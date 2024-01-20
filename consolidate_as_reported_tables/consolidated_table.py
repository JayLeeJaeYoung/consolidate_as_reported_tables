import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz
from itertools import combinations, chain

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, NamedStyle
from colorama import Fore, Back, Style


class Consolidated_Table:  
    """
    - irreconcilable (bool)
    - output_excel_file (str)
    - comp_like_df (DataFrame)
    - manual_mapping_rules (dict)
    - items
    - data (DataFrame): copy of rei.data; used to yield df_long, which produces updated df at the beginning of iteration
    - logger
    - sources_to_consolidate (list): list of sources waiting to be consolidated
    - df_long (DataFrame): long form of df that are already consolidated; used only temporarily to produce updated df at the beginning of iteration
    - df    
    - comp_source (str): current source being consolidated, short for "comparison source"
    - overlapping_periods (list): periods where base and comp overlap
    - comp_only_periods (list): periods only comp applies
    - base_only_periods (list): periods only base applies
    - combination_rules    
    """      
    def __init__(self, rei, output_excel_file, irreconcilable):
        self.irreconcilable = irreconcilable
        self.output_excel_file = output_excel_file
        
        # initialize attributes from rei
        self.comp_like_df = rei.comp_like_df.copy()
        
        if not rei.item_manual_mappings_df.empty:
            if not self.irreconcilable:
                raise ValueError("manual_mapping_rules only applies when irreconcilable is True")
        
            self.manual_mapping_rules = rei.item_manual_mappings_df[['item_from','item_to']].set_index('item_from')['item_to'].to_dict()
           
        self.items = rei.items.copy()
        self.data = rei.data.copy()     
                
        # initialize logger
        self.logger = list()
        
        # prepare to iterate through sources
        self.sources_to_consolidate = rei.metadata_df['tab'].tolist()
        base_source = self.sources_to_consolidate.pop(0)
        self.df_long = self.data[self.data['source'] == base_source].copy()
        self.df_long.loc[:, 'record_type'] = 'base'
        
        self.df = None
        self.comp_source = None
        self.overlapping_periods = None
        self.comp_only_periods = None
        self.base_only_periods = None

        self.combination_rules = []
            
    def _print_df_status(self):
        """
        helper function to print remaining items to be matched
        """
        print(f" -{((~self.df['matched']) & (self.df['record_type'] == 'base') & (self.df['disjoint'] == 'NA')).sum()} items in base to be matched")
        print(f" -{((~self.df['matched']) & (self.df['record_type'] == 'comp') & (self.df['disjoint'] == 'NA')).sum()} items in comp to be matched")    
    
    def prepare_next_source(self):  
        """
        Prepatory step at the beginning of each iteration to prepare updated self.df
        """
        if not self.sources_to_consolidate:
            raise IOError("No more sources to consolidate!")
              
        self.comp_source = self.sources_to_consolidate.pop(0)
        df_new_long = self.data[self.data['source'] == self.comp_source].copy()
        df_new_long.loc[:, 'record_type'] = 'comp'
        self.df_long = pd.concat([self.df_long, df_new_long])
        self.df_long['value'] = self.df_long['value'].fillna(0)
        
        # raw_value is dropped 
        self.df = self.df_long.pivot(
                index=['source', 'record_type','row_num', 'item', 'raw_item'],
                columns='period', values='value') \
            .reset_index() \
            .sort_values(by=['record_type','row_num']) \
            .assign(matched=False, disjoint='NA', init_row_num=lambda x: x['row_num'], init_comp_row_num=-1)
        self.df.columns.name = None

        self.overlapping_periods = list(
            set(self.df_long[self.df_long['record_type'] == 'base']['period'].unique()) & \
            set(self.df_long[self.df_long['record_type'] == 'comp']['period'].unique()))

        self.comp_only_periods = list(
            set(self.df_long[self.df_long['record_type'] == 'comp']['period'].unique()) - \
            set(self.df_long[self.df_long['record_type'] == 'base']['period'].unique())) 
    
        self.base_only_periods = list(
            set(self.df_long[self.df_long['record_type'] == 'base']['period'].unique()) - \
            set(self.df_long[self.df_long['record_type'] == 'comp']['period'].unique())) 
        
        msg = f"start consolidating source={self.comp_source}:"
        print(msg)
        self.logger.append((self.comp_source, msg))
        msg = f"overlapping_periods: {self.overlapping_periods}, comp_only_periods: {self.comp_only_periods}, base_only_periods: {self.base_only_periods}"                
        print(msg)
        self.logger.append((self.comp_source, msg))
            
    def match_same_items(self):
        print('\n1. match_same_items...')
        match_same_item_count = [0]

        def match_same_items_helper(grp):
            if len(grp) == 2:
                if (self.overlapping_periods and grp.duplicated(subset=self.overlapping_periods, keep=False).all()) or (not self.overlapping_periods):
                    grp.loc[grp['record_type'] == 'base', 'init_comp_row_num'] = grp.loc[grp['record_type'] == 'comp', 'init_row_num'].iloc[0]
                    grp.loc[:, 'matched'] = True 
                    # fill in NA
                    grp = grp.ffill()
                    grp = grp.bfill()
                    match_same_item_count[0] += 1
                    
                # if all values in comp are zero, this usually means this iteam signifies some sort of sum or total or sub-total, which no longer shows a value (but shows up as blank)
                # we copy over values from base to comp
                if (grp.loc[grp['record_type'] == 'comp', self.overlapping_periods + self.comp_only_periods] == 0).all(axis=1).all():                    
                    grp.loc[grp['record_type'] == 'base', 'init_comp_row_num'] = grp.loc[grp['record_type'] == 'comp', 'init_row_num'].iloc[0]
                    grp.loc[:, 'matched'] = True 
                    # fill in NA
                    grp = grp.ffill()
                    grp = grp.bfill()
                    # this line is probably unnecessary because we only take base to the next iteration (Note: ffill()/bfill() only fills NA values not zeros)
                    grp.loc[grp['record_type'] == 'comp', self.overlapping_periods + self.comp_only_periods] = grp.loc[grp['record_type'] == 'base', self.overlapping_periods + self.comp_only_periods].values
                    match_same_item_count[0] += 1 
                
            if len(grp) > 2:
                # this should not happen as this has been checked before, but just in case
                raise ValueError(f"Can't have 3 or more items in combined base and comp: {grp.loc[:, 'item'].iloc[0]}")
            return grp
        
        self.df = self.df.groupby('item').apply(match_same_items_helper)
        print(f" {match_same_item_count[0]} pairs matched")
        self._print_df_status()

    def match_same_overlapping_periods_values(self):
        def match_same_overlapping_periods_values_helper(grp):
            if (len(grp) == 2) and (~(grp[self.overlapping_periods] == 0).all(axis=1).all()) and (grp['record_type'].nunique() == 2):
                comp_item_name = grp.loc[grp['record_type'] == 'comp', 'item'].iloc[0]
                comp_init_row_num = grp.loc[grp['record_type'] == 'comp', 'init_row_num'].iloc[0]
                
                base_item_name = grp.loc[grp['record_type'] == 'base', 'item'].iloc[0]
                base_init_row_num = grp.loc[grp['record_type'] == 'base', 'init_row_num'].iloc[0]
                fuzz_ratio = fuzz.ratio(comp_item_name, base_item_name)
                msg = f' item updated with fuzzy ratio of {fuzz_ratio} (row num:{comp_init_row_num:>3}->{base_init_row_num:>3}): {comp_item_name}--> {base_item_name}'
                print(msg)
                self.logger.append((self.comp_source, msg))
                        
                # update self.items
                self.items.loc[(self.items['source'] == self.comp_source) & (self.items['name'] == comp_item_name), 'name'] = base_item_name
                
                # update self.df
                grp.loc[grp['record_type'] == 'base', 'init_comp_row_num'] = comp_init_row_num
                grp.loc[grp['record_type'] == 'comp', 'item'] = base_item_name
                grp.loc[:, 'matched'] = True
                grp = grp.ffill()
                grp = grp.bfill()        
            return grp
        
        print('\n2. match same overlapping periods values...')
        if self.overlapping_periods:            
            unmatched_df = self.df[(~self.df['matched'])]
            unmatched_df = unmatched_df.groupby(self.overlapping_periods).apply(match_same_overlapping_periods_values_helper)
            self.df.loc[unmatched_df.index] = unmatched_df
        self._print_df_status()

    def manually_map_items(self):
        if not self.irreconcilable:
            return
        
        print('\n3. manually map inconsistent items...')
        
        for item_from, item_to in self.manual_mapping_rules.items():
            # Note: (~self.df['matched']) is not a condition
            is_item_from_row = (self.df['record_type'] == 'comp') & (self.df['item'] == item_from)
            is_item_to_row = (self.df['record_type'] == 'base') & (self.df['item'] == item_to)
            
            if not is_item_from_row.any() or not is_item_to_row.any(): 
                continue
                    
            # we don't check whether values of overlapping_periods are consistent because it won't be.
            
            # update self.items
            self.items.loc[(self.items['source'] == self.comp_source) & (self.items['name'] == item_from), 'name'] = item_to

            # update self.df
            # fill NA values in base from comp
            self.df.loc[is_item_to_row, self.comp_only_periods] = self.df.loc[is_item_from_row, self.comp_only_periods].sum().values

            # writes the min init_row_num to init_comp_row_num
            self.df.loc[is_item_to_row, 'init_comp_row_num'] = self.df.loc[is_item_from_row, 'init_row_num'].min()
            
            self.df.loc[is_item_from_row, 'item'] = self.df.loc[is_item_to_row, 'item'].values
            
            # self.df.loc[is_item_to_row, 'matched'] = True
            # self.df.loc[is_item_from_row, 'matched'] = True   
            msg = f" manually map inconsistent items applied: {item_from} --> {item_to}"
            print(msg)
            self.logger.append((self.comp_source, msg))
    
        self._print_df_status()

    def apply_combination_rules(self):
        print('\n4. apply combination rules...')
                        
        for x in self.combination_rules:
            base_tuple_mask = (self.df['item'].isin(x['base_tuple'])) & (self.df['record_type']=='base')
            comp_tuple_mask = (self.df['item'].isin(x['comp_tuple'])) & (self.df['record_type']=='base')
            
            # since we are keeping track of duplicated items, we expect none or 1 items to be matched by this step, but not both
            if self.df[comp_tuple_mask]['matched'].all() and self.df[base_tuple_mask]['matched'].all():
                raise ValueError(f"both comp_tuple: {x['comp_tuple']} and base_tuple: {x['base_tuple']} are already matched")
                                      
            # case when base_tuple is 1 and comp_tuple is more than 1
            if base_tuple_mask.sum() == 1:
                # comp_tuple is matched in previous steps
                if self.df[comp_tuple_mask]['matched'].all():
                    if (self.df.loc[base_tuple_mask, self.overlapping_periods].sum().values == self.df.loc[comp_tuple_mask, self.overlapping_periods].sum().values).all():
                        self.df.loc[base_tuple_mask, self.comp_only_periods] =  self.df.loc[comp_tuple_mask, self.comp_only_periods].sum().values
                        x['sources'].append(self.comp_source) 
                        self.df.loc[base_tuple_mask, 'matched'] = True
                        msg = f'rule applied and {x["base_tuple"]} copied over: {x}'
                    else:
                        x['invalid_sources'].append(self.comp_source)
                        msg = f'rule invalid: {x}'
                        
                # base_tuple is matched in previous steps
                elif self.df[base_tuple_mask]['matched'].all():
                    self.df.loc[comp_tuple_mask, 'matched'] = True 
                    x['sources'].append(self.comp_source) 
                    msg = f'rule applied: {x}'
                else:
                    msg = f'rule irrelevant: {x}'
                    # self.df.loc[base_tuple_mask, 'matched'] = True 
                    # self.df.loc[comp_tuple_mask, 'matched'] = True 
                    
            # case when base_tuple is more than 1 and comp_tuple is 1
            elif comp_tuple_mask.sum() == 1:
                # comp_tuple is matched in previous steps
                if self.df[comp_tuple_mask]['matched'].all():
                    self.df.loc[base_tuple_mask, 'matched'] = True    
                    x['sources'].append(self.comp_source)
                    msg = f'rule applied: {x}'
                     
                # base_tuple is matched in previous steps
                elif self.df[base_tuple_mask]['matched'].all():
                    if (self.df.loc[base_tuple_mask, self.overlapping_periods].sum().values == self.df.loc[comp_tuple_mask, self.overlapping_periods].sum().values).all():
                        self.df.loc[comp_tuple_mask, self.comp_only_periods] =  self.df.loc[base_tuple_mask, self.comp_only_periods].sum().values
                        x['sources'].append(self.comp_source) 
                        self.df.loc[comp_tuple_mask, 'matched'] = True 
                        msg = f'rule applied and {x["comp_tuple"]} copied over: {x}'
                    else:
                        x['invalid_sources'].append(self.comp_source)
                        msg = f'rule invalid: {x}'
                else:
                    msg = f'rule irrelevant: {x}'
                    # self.df.loc[base_tuple_mask, 'matched'] = True 
                    # self.df.loc[comp_tuple_mask, 'matched'] = True               
            else:
                raise ValueError("M:M mapping between comp_tuples and base_tuples not allowed")            
            print(msg)
            self.logger.append((self.comp_source, msg))
    
    def designate_disjoint_items(self):
        print('\n6a. set aside disjoint items...')

        mask = (~self.df['matched'] & (self.df['record_type'] == 'base') & ((self.df[self.overlapping_periods] == 0).all(axis=1)))
        print(f" {mask.sum()} items designated as disjoint base")
        self.df.loc[mask, 'disjoint'] = 'base'
        mask = (~self.df['matched'] & (self.df['record_type'] == 'comp') & ((self.df[self.overlapping_periods] == 0).all(axis=1)))
        print(f" {mask.sum()} items designated as disjoint comp")
        self.df.loc[mask, 'disjoint'] = 'comp'

        comp_like_series = self.comp_like_df[self.comp_like_df['source'] == self.comp_source]['raw_item']
        mask = self.df['raw_item'].isin(comp_like_series)
        print(f" {mask.sum()} items designated as disjoint comp_like")
        self.df.loc[mask, 'disjoint'] = 'comp_like'
        self._print_df_status()
        
    def _manually_reconcile(self, unmatched_base_list):
        # ignore comp and just use values in base
        for item in unmatched_base_list:
            self.df.loc[(self.df['item'] == item) & (self.df['record_type']=='base') , self.comp_only_periods] = self.df.loc[(self.df['item'] == item) & (self.df['record_type']=='comp') , self.comp_only_periods].values
            self.df.loc[(self.df['item'] == item) & (self.df['record_type']=='base') , 'init_comp_row_num'   ] = self.df.loc[(self.df['item'] == item) & (self.df['record_type']=='comp') , 'init_row_num'        ].min()
            self.df.loc[self.df['item'] == item , 'matched'] = True
    
    def apply_combinations_to_match(self):
        print('\n5. apply combinations to match...')
        
        unmatched_base_mask = (~self.df['matched']) & (self.df['disjoint'] =='NA') & (self.df['record_type']=='base')
        unmatched_comp_mask = (~self.df['matched']) & (self.df['disjoint'] =='NA') & (self.df['record_type']=='comp')
        unmatched_base_list = self.df[unmatched_base_mask]['item'].tolist()
        unmatched_comp_list = self.df[unmatched_comp_mask]['item'].tolist()

        print(f"unmatched_base_list: {unmatched_base_list}")
        print(f"unmatched_comp_list: {unmatched_comp_list}")
    
        # check at least the sum of the unmatched items match in base and comp
        # it's OK to have "waste" comp rows that are non-zero as long as base for overlapping period is zero...there are cases it is impossible to reconcile
        if unmatched_base_list:
            if (self.df[unmatched_base_mask][self.overlapping_periods].sum() == self.df[unmatched_comp_mask][self.overlapping_periods].sum()).all():
                pass
            else:
                if tuple(sorted(unmatched_base_list)) == tuple(sorted(unmatched_comp_list)):
                    msg = f'Inconsistent data:\nunmatched_base_list: {unmatched_base_list}\nunmatched_comp_list: {unmatched_comp_list}'
                    # warnings.warn(msg)
                    print(f"{Fore.BLACK}{Back.MAGENTA}{Style.BRIGHT}{msg}{Style.RESET_ALL}")
                    self.logger.append((self.comp_source, msg))
                    if self.irreconcilable:
                        self._manually_reconcile(unmatched_base_list)
                    else:
                        raise ValueError("stopping operation due to inconsistent data")
                    self._print_df_status()
                    return
                elif (len(set(unmatched_comp_list) - set(unmatched_base_list)) == 0):
                    disappeared_items = set(unmatched_base_list) - set(unmatched_comp_list)
                    msg = f'Inconsistent data:\nunmatched_base_list: {unmatched_base_list}\nunmatched_comp_list: {unmatched_comp_list}\nfrom unmatched_base_list {disappeared_items} disappeared in comp'
                    print(f"{Fore.BLACK}{Back.MAGENTA}{Style.BRIGHT}{msg}{Style.RESET_ALL}")
                    self.logger.append((self.comp_source, msg))
                    if self.irreconcilable:
                        # we ignore items in base that no longer appear in comp
                        self.df.loc[self.df['item'].isin(disappeared_items), 'disjoint'] = 'base'
                        self._manually_reconcile(unmatched_comp_list)
                    else:
                        raise ValueError("stopping operation due to inconsistent data")         
                    self._print_df_status()
                    return
                else:
                    msg = f"Inconsistent data:\nOnly in unmatched base: {set(unmatched_base_list) - set(unmatched_comp_list)}\nOnly in unmatched comp: {set(unmatched_comp_list) - set(unmatched_base_list)}"
                    raise ValueError(msg)

        compatiable_comp_items = {k: [] for k in unmatched_base_list}
        for base_item in unmatched_base_list:
            # print(f"look at base_item: {base_item}")
            base_row = self.df[(self.df['item'] == base_item) & (self.df['record_type'] == 'base')]
            for x in chain.from_iterable(combinations(unmatched_comp_list, r) for r in range(2, len(unmatched_comp_list) + 1)):
                comp_rows = self.df[(self.df['item'].isin(x)) & (self.df['record_type'] == 'comp')]
                if (base_row[self.overlapping_periods].sum() == comp_rows[self.overlapping_periods].sum()).all():
                    compatiable_comp_items.setdefault(base_item, []).append(x)
                    break # on finding the first match, we break
                
                
        # check compatiable_comp_items and add to new_rules
        for base_item, comp_items in compatiable_comp_items.items():
            if len(comp_items) == 1:
                msg = f' new rule created: base {base_item} -> comps {comp_items[0]}'
                print(msg)
                self.logger.append((self.comp_source, msg))
            elif len(comp_items) == 0:
                msg = f"There is unmatched base item: {base_item}"
                print(f"{Fore.BLACK}{Back.MAGENTA}{Style.BRIGHT}{msg}{Style.RESET_ALL}")
                # raise ValueError("There is unmatched base item")
            else:
                # this case won't be triggered because I now put break when I find the first match
                raise ValueError("There are multiple possible matchings")
    
        for base_item, comp_items in compatiable_comp_items.items():
            if len(comp_items) != 1:
                continue
            
            # update self.items
            new_comp_item_series = self.items.loc[(self.items['source'] == self.comp_source) & (self.items['name'].isin(comp_items[0])), 'name'] + ' <<1>>' 
            self.items.loc[(self.items['source'] == self.comp_source) & (self.items['name'].isin(comp_items[0])), 'name'] = new_comp_item_series
                
            # update self.df
            is_base_row = (self.df['item'] == base_item) & (self.df['record_type'] == 'base')
            is_comp_row = (self.df['item'].isin(comp_items[0])) & (self.df['record_type'] == 'comp')
            
            self.df.loc[is_comp_row, 'item'] = self.df.loc[is_comp_row, 'item'] + ' <<1>>'
            # self.df.loc[is_comp_row, 'matched'] = True
            self.df.loc[is_comp_row, 'disjoint'] = 'comp_like'
            
            self.df.loc[is_base_row, self.comp_only_periods] = self.df.loc[is_comp_row, self.comp_only_periods].sum().values
            self.df.loc[is_base_row, 'init_comp_row_num'] = self.df.loc[is_comp_row, 'init_row_num'].min()   
            self.df.loc[is_base_row, 'matched'] = True   
            
            self.combination_rules.append({
                'comp_tuple': tuple(self.df.loc[is_comp_row, 'item']),
                'base_tuple': tuple(self.df.loc[is_base_row, 'item']),
                'sources': [self.comp_source],
                'invalid_sources': []
            })
            
        unmatched_base_mask = (~self.df['matched']) & (self.df['disjoint'] =='NA') & (self.df['record_type']=='base')
        unmatched_comp_mask = (~self.df['matched']) & (self.df['disjoint'] =='NA') & (self.df['record_type']=='comp')
        unmatched_base_list = self.df[unmatched_base_mask]['item'].tolist()
        unmatched_comp_list = self.df[unmatched_comp_mask]['item'].tolist()
        
          
        compatiable_base_items = {k: [] for k in unmatched_comp_list}
        for comp_item in unmatched_comp_list:
            comp_row = self.df[(self.df['item'] == comp_item) & (self.df['record_type'] == 'comp')]
            for x in chain.from_iterable(combinations(unmatched_base_list, r) for r in range(2, len(unmatched_base_list) + 1)):
                base_rows = self.df[(self.df['item'].isin(x)) & (self.df['record_type'] == 'base')]
                if (comp_row[self.overlapping_periods].sum() == base_rows[self.overlapping_periods].sum()).all():
                    compatiable_base_items.setdefault(comp_item, []).append(x)
                    break # on finding the first match, we break
                
        # check compatiable_base_items and add to new_rules
        for comp_item, base_items in compatiable_base_items.items():
            print(comp_item, base_items)
            if len(base_items) == 1:
                msg = f' new rule created: comp {comp_item} -> bases {base_items[0]}'
                print(msg)
                self.logger.append((self.comp_source, msg))
            elif len(base_items) == 0:
                msg = f"There is unmatched comp item: {comp_item}"
                raise ValueError(msg)
            else:
                # this case won't be triggered because I now put break when I find the first match
                raise ValueError("There are multiple possible matchings")
                    
           
        for comp_item, base_items in compatiable_base_items.items():
            
            # update self.items            
            new_comp_item = self.items.loc[(self.items['source'] == self.comp_source) & (self.items['name'] == comp_item), 'name'] + ' [[1]]' 
            self.items.loc[(self.items['source'] == self.comp_source) & (self.items['name'] == comp_item), 'name'] = new_comp_item
                
            # update self.df
            is_base_row = (self.df['item'].isin(base_items[0])) & (self.df['record_type'] == 'base')
            is_comp_row = (self.df['item'] == comp_item) & (self.df['record_type'] == 'comp')
            
            self.df.loc[is_comp_row, 'item'] = self.df.loc[is_comp_row, 'item'] + ' [[1]]'
            self.df.loc[is_comp_row, self.base_only_periods] = self.df.loc[is_base_row, self.base_only_periods].sum().values
            # self.df.loc[is_comp_row, 'matched'] = True
            self.df.loc[is_comp_row, 'disjoint'] = 'comp_like'
            
            
            self.df.loc[is_base_row, 'init_comp_row_num'] = self.df.loc[is_comp_row, 'init_row_num'].min()   
            self.df.loc[is_base_row, 'matched'] = True   
            
            self.combination_rules.append({
                'comp_tuple': tuple(self.df.loc[is_comp_row, 'item']),
                'base_tuple': tuple(self.df.loc[is_base_row, 'item']),
                'sources': [self.comp_source],
                'invalid_sources': []
            })     
            
            
        # Before embracing "disjoint" rows, must check there are not leaks
        if unmatched_base_list:
            if not (self.df[unmatched_base_mask][self.overlapping_periods].sum() == self.df[unmatched_comp_mask][self.overlapping_periods].sum()).all():
                raise ValueError('leaks?')
        self._print_df_status()
    
    def apply_disjoint_items(self):
        print('\n6b. apply disjoint items...')
        
        mask = self.df['disjoint'] == 'base'
        self.df.loc[mask, 'matched'] = True        
        print(f" {mask.sum()} disjoint base items matched")
        
        
        # """    
        # The rows with all values = 0 for overlapping_periods are rows that we cannot verify.
        # The rows either exist only in base or in comp. Let's take those among comp.
        # --> We take each such comp row and insert them into 'appropriate' row positions in base.
        # We will create new rows in base base and 'matched'
        # """

        # must be reverse order to avoid nan
        comp_item_init_row_num_list = np.sort(self.df.loc[self.df['disjoint'].isin(['comp', 'comp_like']), 'init_row_num'].values).tolist()
        for comp_item_init_row_num in comp_item_init_row_num_list:
            # find the last comp row that got converted into base that is smaller than the this comp row    
            last_init_comp_row_num =  self.df[(self.df['record_type'] == 'base') & (self.df['init_comp_row_num'] > 0) & \
                (self.df['init_comp_row_num'] <= comp_item_init_row_num)]['init_comp_row_num'].max()                
            insert_min_row_num = self.df[self.df['init_comp_row_num'] == last_init_comp_row_num]['row_num'].iloc[0]
            print(f' insert_min_row_num: {insert_min_row_num} for comp_item_init_row_num={comp_item_init_row_num}')
            
            
            # for all the row_num greater than insert_min_row_num, increment by one
            self.df.loc[(self.df['record_type'] == 'base') & (self.df['row_num'] > insert_min_row_num), 'row_num'] += 1
            
            # fill in that new empty item with comp --> base
            new_base_row = self.df.loc[(self.df['record_type'] == 'comp') & (self.df['init_row_num'] == comp_item_init_row_num)].copy()
            new_base_row.loc[:, 'row_num'] = insert_min_row_num + 1    
            new_base_row.loc[:, 'init_comp_row_num'] =  new_base_row.loc[:, 'init_row_num'].iloc[0]
            new_base_row.loc[:, 'matched'] = True 
            new_base_row.loc[:, 'record_type'] = 'base'
            self.df.loc[(self.df['record_type'] == 'comp') & (self.df['init_row_num'] == comp_item_init_row_num), 'matched'] = True    
            self.df = pd.concat([self.df, new_base_row], ignore_index=True)
    
        
        if ((~self.df['matched']) & (self.df['disjoint'] != 'NA')).sum() > 0:
            raise ValueError(f"There are still {((~self.df['matched']) & (self.df['disjoint'] != 'NA')).sum()} disjoint items left")
        self._print_df_status()
       
    def _export_consolidated_table(self):
        # new_rules = self._get_new_rules()
        # for comp_tuples, base_tuples in new_rules.items():
        #     self.logger.append(('Final', f"New Rule: {comp_tuples} --> {base_tuples}"))
    
        with pd.ExcelWriter(self.output_excel_file) as writer:
            self.df.drop(columns=['record_type']).to_excel(writer, sheet_name='table', index=False)
            self.items.to_excel(writer, sheet_name='items', index=False)
            pd.DataFrame(self.logger).to_excel(writer, sheet_name='log', index=False, header=False)
        print(f'Finished! Results are exported to {self.output_excel_file}.')
        
    def post_process_next_source(self):    
        print(f"\nwrapping up consolidating source={self.comp_source}...")
        
        
        if (~self.df['matched']).sum() != 0:
            raise ValueError("There are still unmatched rows...")
        
        del self.df['matched']
        del self.df['init_row_num']
        del self.df['init_comp_row_num']
        del self.df['disjoint']
    
        self.df = self.df.sort_values(by=['record_type','row_num'])
        self.df = self.df[self.df['record_type'] == 'base'].copy()
        self.df_long = self.df.melt(
                id_vars=['source', 'record_type', 'row_num', 'item', 'raw_item'],
                var_name='period',
                value_name='value') \
            .assign(raw_value=lambda x: x['value'])[['source', 'record_type', 'period', 'row_num', 'item', 'raw_item', 'value', 'raw_value']]
        self.df_long['value'] = self.df_long['value'].fillna(0)
        
        # if there is no more source to consolidate, export results and return
        if not self.sources_to_consolidate:
            self._export_consolidated_table()
            return
        print(self.sources_to_consolidate)
        print("Done.")

    def consolidate_next_source(self):
        self.prepare_next_source()
        self.match_same_items()
        self.match_same_overlapping_periods_values()
        self.manually_map_items()
        self.apply_combination_rules()
        self.designate_disjoint_items()
        self.apply_combinations_to_match()
        self.apply_disjoint_items()
        self.post_process_next_source()

    def debug_export_df(self, debug_file_name):     
        """
        export current df for debugging purpose
        """
        pre_columns = ['source', 'record_type', 'row_num' ,'item', 'raw_item']
        period_columns = sorted(self.comp_only_periods + self.overlapping_periods)
        post_columns = ['matched', 'disjoint', 'init_row_num', 'init_comp_row_num']
        display_columns = pre_columns  + period_columns + post_columns

        df_export = self.df[display_columns].copy()

        # Create 'base' and 'comp' sheets
        with pd.ExcelWriter(debug_file_name) as writer:
            base_df = df_export.loc[self.df['record_type']=='base']
            comp_df = df_export.loc[self.df['record_type']=='comp']
            base_df.to_excel(writer, sheet_name='base', index=False)
            comp_df.to_excel(writer, sheet_name='comp', index=False)

        # Load the Excel file using openpyxl
        workbook = load_workbook(debug_file_name)

        # Create a custom number format style
        custom_number_format = NamedStyle(name='custom_number_format')
        custom_number_format.number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-' # 2 decimal
        custom_number_format.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-' # 0 decimal

        # Define fill styles for highlighting
        highlight_fill_matched = PatternFill(start_color="F3F549", end_color="F3F549", fill_type="solid")  # Yellow for matched=False
        highlight_fill_disjoint = PatternFill(start_color="E0DDDC", end_color="E0DDDC", fill_type="solid")  # Grey for disjoint

        for sheet_name in ['base', 'comp']:
            sheet = workbook[sheet_name]
            for col in sheet.iter_cols(max_col=100):
                # if the first row's value contains period_columns
                if col[0].value in period_columns:
                    # pick rows after 2nd row
                    for cell in col[1:]:
                        cell.style = custom_number_format
                        
            for row in sheet.iter_rows(min_row=2, max_col=len(display_columns)):
                # if the 'matched' column is False, highlight
                if row[display_columns.index('matched')].value == False:
                    for cell in row:
                        cell.fill = highlight_fill_matched
                        
                # if the 'disjoint' column is not NA, highlight
                if row[display_columns.index('disjoint')].value != 'NA':
                    for cell in row:
                        cell.fill = highlight_fill_disjoint        


        # create 'diff' sheet             
        diff_sheet = workbook.create_sheet('diff')

        # copy 'base' sheet content to 'diff' sheet
        sheet = workbook['base']
        for row in sheet.iter_rows(min_col=4, max_col=len(display_columns)):
            for cell in row:
                diff_cell = diff_sheet.cell(row=cell.row, column=cell.column - 3, value=cell.value);
                diff_cell._style = cell._style

        # copy 'comp' sheet content to 'diff' sheet
        sheet = workbook['comp']        
        for row in sheet.iter_rows(min_col=4, max_col=len(display_columns)):
            for cell in row:
                diff_cell = diff_sheet.cell(row=cell.row, column=cell.column - 3 + len(display_columns) - 2, value=cell.value);
                diff_cell._style = cell._style

        # Save the modified workbook with the desired filename
        workbook.save(debug_file_name)      
        